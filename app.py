from flask import Flask, render_template, request, redirect, url_for, session, jsonify
from openpyxl import load_workbook
import sqlite3
import datetime

app = Flask(__name__)
app.secret_key = "your_secret_key_here"


@app.context_processor
def common_icons():
    admin = "/static/assets/admin.png"
    student = "/static/assets/student.png"
    faculty = "/static/assets/faculty.png"
    search = "/static/assets/search.png"
    notification = "/static/assets/notification.png"
    settings = "/static/assets/settings.png"
    return dict(
        search=search,
        notification=notification,
        settings=settings,
        admin=admin,
        faculty=faculty,
        student=student,
    )


@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        user_type_form = request.form.get("user-type")
        username = request.form.get("username")
        password = request.form.get("password")

        conn = sqlite3.connect("database.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, role, username FROM users WHERE username=? AND password=?",
            (username, password),
        )
        result = cursor.fetchone()

        if result is not None:
            user_id, user_type, name = result
            session["user_id"] = user_id
            session["user_type"] = user_type
            session["username"] = username
            session["name"] = name

            if user_type_form == user_type:
                return redirect(url_for("home"))
        error_message = "Invalid username or password"
        return render_template("error.html", error=error_message)

    return render_template("login.html")


@app.route("/home", endpoint="home")
def home():
    user_id = session.get("user_id")
    user_type = session.get("user_type")
    username = session.get("username")
    name = session.get("name")
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    if user_type == "student":
        cursor.execute("SELECT name FROM studentlist WHERE id=?", (user_id,))
    elif user_type == "faculty":
        cursor.execute("SELECT name FROM facultylist WHERE id=?", (user_id,))

    result = cursor.fetchone()
    fullname = None
    if result is not None:
        fullname = result[0]

    session["fullname"] = fullname
    cursor.execute("SELECT status FROM attendance WHERE id=?", (user_id,))
    attendance_data = [status[0] for status in cursor.fetchall()]

    total_attendance = len(attendance_data)
    present_attendance = sum(1 for status in attendance_data if status == "Present")

    attendance_percentage = 0
    if total_attendance != 0:
        attendance_percentage = (present_attendance / total_attendance) * 100

    conn.close()

    return render_template(
        "home.html",
        username=username,
        name=name,
        user_type=user_type,
        fullname=fullname,
        attendance_percentage=attendance_percentage,
    )


@app.route("/studentlist")
def get_student_list():
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("SELECT id, name ,email FROM studentlist")
    student_list = c.fetchall()
    conn.close()
    return jsonify(student_list)


@app.route("/attendancepercentage")
def attendancepercentage():
    user_id = session.get("user_id")
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute("SELECT SUM(status) FROM attendance WHERE id=?", (user_id,))
    sum_status = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM attendance WHERE id=?", (user_id,))
    num_rows = cursor.fetchone()[0]

    attendance_percentage = (sum_status / num_rows) * 100 if num_rows > 0 else 0
    attendance_percentage = round(attendance_percentage, 1)

    cursor.close()
    conn.close()

    return jsonify(attendance_percentage)


@app.route("/grades")
def get_student_grades():
    user_id = session.get("user_id")
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("SELECT s1, s2, s3, s4, s5, s6 FROM grades WHERE id = ?", (user_id,))
    data = c.fetchone()
    conn.close()
    if data:
        grade_data = {
            "s1": data[0],
            "s2": data[1],
            "s3": data[2],
            "s4": data[3],
            "s5": data[4],
            "s6": data[5],
        }
        return jsonify(grade_data)
    else:
        return jsonify({})


@app.route("/tasks")
def get_tasks():
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("SELECT name, due, status FROM tasks")
    data = c.fetchall()
    conn.close()

    tasks = []
    for task in data:
        task_dict = {"name": task[0], "due": task[1], "status": task[2]}
        tasks.append(task_dict)

    return jsonify(tasks)


@app.route("/perfomance")
def perfomance():
    user_id = session.get("user_id")
    user_type = session.get("user_type")
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM grade WHERE id = ?", (user_id,))
    marks = cursor.fetchone()
    
    cursor.execute("SELECT * FROM grades")
    grades = cursor.fetchall()
    conn.close()
    return render_template("perfomance.html", user_type=user_type, marks=marks,grades=grades)


@app.route("/profile")
def profile():
    user_id = session.get("user_id")
    user_type = session.get("user_type")
    fullname = session.get("fullname")

    if user_id is None:
        return "User ID not found"

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    if user_type == "student":
        cursor.execute("SELECT * FROM studentlist WHERE id=?", (user_id,))
        student_info = cursor.fetchone()
        conn.close()
        

        if student_info is not None:
            id = student_info[0]
            name = student_info[1]
            dob = student_info[2]
            email = student_info[3]
            course = student_info[4]
            contact = student_info[5]
            address = student_info[6]

            return render_template(
                "profile.html",
                id=id,
                user_type=user_type,
                name=name,
                dob=dob,
                email=email,
                fullname=fullname,
                course=course,
                contact=contact,
                address=address,
            )

    elif user_type == "faculty":
        cursor.execute("SELECT * FROM facultylist WHERE id=?", (user_id,))
        faculty_info = cursor.fetchone()
        conn.close()

        if faculty_info is not None:
            id = faculty_info[0]
            name = faculty_info[1]
            dob = faculty_info[2]
            email = faculty_info[3]
            department = faculty_info[4]
            contact = faculty_info[5]
            address = faculty_info[6]

            return render_template(
                "profile.html",
                id=id,
                user_type=user_type,
                dob=dob,
                course=department,
                name=name,
                email=email,
                contact=contact,
                address=address,
                fullname=fullname,
            )

    return "User not found"


@app.route("/upload")
def upload():
    return render_template("upload-faculty.html")


@app.route("/process", methods=["POST"])
def process():
    file = request.files["file"]
    wb = load_workbook(file)
    sheet = wb.active
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS attendance (
            date,
            student_id,
            attendance_status
        )
    """
    )
    for row in sheet.iter_rows(min_row=2):
        date, student_id, attendance_status = row
        cursor.execute(
            "INSERT INTO attendance VALUES (?, ?, ?)",
            (date, student_id, attendance_status),
        )
    conn.commit()
    conn.close()
    message = "File uploaded successfully!"
    return render_template("upload.html", message=message)


@app.route("/attendance", methods=["GET", "POST"])
def attendance():
    user_id = session.get("user_id")
    user_type = session.get("user_type")

    if request.method == "POST":
        selected_month = request.form["month"]
    else:
        selected_month = "January"  # Default month

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute(
        "SELECT status FROM attendance WHERE id=? AND month=?",
        (user_id, selected_month),
    )
    attendance_records = cursor.fetchall()

    conn.close()

    if len(attendance_records) == 0:
        return f"No attendance records found for user ID {user_id} in {selected_month}."

    return render_template(
        "attendance.html",
        user_type=user_type,
        attendance_records=attendance_records,
        selected_month=selected_month,
    )


@app.route("/classAttendance", methods=["GET", "POST"])
def classAttendance():
    user_type = session.get("user_type")
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    selected_day = request.args.get("day")
    if selected_day is None:
        selected_day = "1"

    cursor.execute("SELECT DISTINCT date FROM attendance")
    dates = cursor.fetchall()
    days = [str(date[0]) for date in dates]

    cursor.execute(
        "SELECT id, status FROM attendance WHERE date=?",
        (selected_day,),
    )
    rows = cursor.fetchall()

    attendance_data = {}
    for row in rows:
        student_id = row[0]
        attendance_status = row[1]

        attendance_data[student_id] = attendance_status

    conn.close()
    return render_template(
        "attendance.html",
        user_type=user_type,
        attendance_data=attendance_data,
        days=days,
        selected_day=selected_day,
    )


@app.route("/courses")
def courses():
    user_type = session.get("user_type")
    fullname = session.get("fullname")
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM courses")
    courses_data = cursor.fetchall()
    conn.close()
    return render_template("courses.html", courses=courses_data, user_type=user_type,fullname=fullname,)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


if __name__ == "__main__":
    app.run(debug=True)
